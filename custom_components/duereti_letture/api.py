"""Client per le API pubbliche Duereti (estrazione curve/letture)."""
from __future__ import annotations

import asyncio
import base64
import csv
import io
import logging
import re
import time
import zipfile
from dataclasses import dataclass, field
from datetime import date, timedelta

import aiohttp

from .const import (
    ESITO_OK,
    MAX_DATE_RANGE_MONTHS,
    MAX_SUPPLY_POINTS_PER_REQUEST,
    MODE_CURVE,
    RESULT_POLL_INTERVAL_SECONDS,
    RESULT_POLL_MAX_ATTEMPTS,
    RIAVVIO_RETRY_ATTEMPTS,
    RIAVVIO_RETRY_WAIT_SECONDS,
    TOKEN_SAFETY_MARGIN_SECONDS,
    TOKEN_VALIDITY_SECONDS,
    URL_REQUEST_EXPORT,
    URL_REQUEST_RESULT,
    URL_REQUEST_TOKEN,
)

_LOGGER = logging.getLogger(__name__)

# Chiavi il cui valore non va MAI loggato in chiaro: le credenziali API e il
# token di sessione. home-assistant.log viene spesso condiviso per chiedere
# supporto, quindi qualunque cosa finisca qui va considerata pubblica.
_CHIAVI_SEGRETE = {"clientid", "secretid", "authorization", "token"}


def _oscura(valore: object) -> str:
    """Riduce un segreto a una forma che resta utile per il debug (si capisce
    se è presente, se cambia tra una chiamata e l'altra, quanto è lungo) senza
    rivelarlo."""
    testo = str(valore)
    if not testo:
        return "<vuoto>"
    if len(testo) <= 8:
        return f"<{len(testo)} caratteri>"
    return f"{testo[:4]}…{testo[-2:]} (<{len(testo)} caratteri>)"


def _sanitizza(dati: dict | None) -> dict:
    """Copia di un dict con i valori sensibili oscurati, pronta per il log."""
    if not dati:
        return {}
    return {
        chiave: (_oscura(valore) if chiave.lower() in _CHIAVI_SEGRETE else valore)
        for chiave, valore in dati.items()
    }


def descrivi_errore(err: Exception) -> str:
    """Rende leggibile un errore delle API, in particolare il blocco del WAF.

    Quando il WAF (F5 BIG-IP) rifiuta una richiesta restituisce una pagina
    HTML al posto del JSON: mostrarla intera all'utente è inutile e
    incomprensibile. Ne estraiamo il solo Support ID, che è l'unica
    informazione utile (serve a Duereti per rintracciare il blocco nei log).
    """
    testo = str(err)
    if "Request Rejected" not in testo:
        return testo
    match = re.search(r"support ID is:\s*(\d+)", testo, re.IGNORECASE)
    support_id = match.group(1) if match else "non disponibile"
    return (
        "richiesta rifiutata dal WAF di Duereti (blocco intermittente, non dipende "
        f"dalle credenziali). Support ID: {support_id}"
    )


def _log_richiesta(url: str, body: dict | None, headers: dict | None) -> None:
    """Logga la richiesta in uscita (solo a livello DEBUG).

    Serve a confrontare byte per byte cosa manda l'integrazione rispetto a un
    client che funziona (curl/Bruno), utile per capire perché il WAF di
    Duereti rifiuta le une e non le altre.
    """
    if not _LOGGER.isEnabledFor(logging.DEBUG):
        return
    _LOGGER.debug(
        "--> POST %s\n    headers: %s\n    body: %s",
        url,
        _sanitizza(headers),
        _sanitizza(body),
    )


def _log_risposta(url: str, status: int, headers: object, corpo: str) -> None:
    """Logga la risposta grezza (solo a livello DEBUG).

    Il corpo viene loggato così com'è (troncato): quando il WAF interviene non
    è JSON ma una pagina HTML "Request Rejected", e vedere il testo esatto con
    il Support ID è proprio l'informazione che serve.
    """
    if not _LOGGER.isEnabledFor(logging.DEBUG):
        return
    anteprima = corpo if len(corpo) <= 800 else f"{corpo[:800]}… [troncato, {len(corpo)} caratteri]"
    _LOGGER.debug(
        "<-- %s da %s\n    headers: %s\n    body: %s",
        status,
        url,
        dict(headers) if headers else {},
        anteprima,
    )


class DuretiApiError(Exception):
    """Errore generico nella comunicazione con le API Duereti."""

    def __init__(self, message: str, http_status: int | None = None, data: dict | None = None) -> None:
        super().__init__(message)
        self.http_status = http_status
        self.data = data or {}


class DuretiAuthError(DuretiApiError):
    """401 UNAUTHORIZED - credenziali o token non validi/scaduti."""


class DuretiValidationError(DuretiApiError):
    """400 BAD_REQUEST - validazione input fallita (dettagli in message)."""


class DuretiConflictError(DuretiApiError):
    """409 CONFLICT - token già esistente per l'utente (requestToken)."""


class DuretiRateLimitError(DuretiApiError):
    """429 - limite di richieste attive raggiunto."""


class DuretiNotFoundError(DuretiApiError):
    """404 NOT_FOUND - ticket presente ma non collegato a nessun dato."""


@dataclass
class CurvaPunto:
    """Un singolo punto curva (timestamp + valore)."""

    timestamp: "datetime"  # noqa: F821 - import sotto per evitare cicli
    valore_kwh: float
    # Valore grezzo della colonna FL_ORA_LEGALE del CSV. Serve a disambiguare
    # l'ora ripetuta al passaggio da ora legale a ora solare: il timestamp da
    # solo è ambiguo, questo flag dice a quale delle due ore appartiene.
    # Vedi statistics._offset_da_flag per l'interpretazione.
    ora_legale: str | None = None


@dataclass
class RisultatoLetture:
    """Risultato del parsing dello zip per un singolo POD."""

    pod: str
    punti: list = field(default_factory=list)  # list[CurvaPunto]


class DuretiApiClient:
    """Wrapper asincrono per il flusso requestToken -> requestExport -> requestResult."""

    def __init__(self, session: aiohttp.ClientSession, client_id: str, secret_id: str) -> None:
        self._session = session
        self._client_id = client_id
        self._secret_id = secret_id
        self._token: str | None = None
        self._token_expiry: float = 0.0

    async def async_assicura_token(self) -> None:
        """Si assicura di avere un token valido, riusando quello in cache.

        A differenza di async_validate_credentials non forza il rinnovo: serve
        al coordinator per sapere se la prima chiamata (requestToken) funziona,
        senza sprecare richieste verso un'API che le blocca a intermittenza.
        """
        await self._get_token()

    async def async_validate_credentials(self) -> None:
        """Verifica che Client ID/Secret ID siano validi.

        Solleva DuretiAuthError se le credenziali sono sbagliate, o
        DuretiApiError per qualunque altro problema di connessione. Non
        restituisce nulla se la validazione ha successo. Metodo pubblico
        pensato per config_flow.py (evita di dover chiamare _get_token,
        un metodo interno, da fuori la classe).
        """
        await self._get_token(force_refresh=True)

    async def _get_token(self, force_refresh: bool = False) -> str:
        """Ottiene un token valido, richiedendone uno nuovo se scaduto o mancante.

        Il manuale documenta un 409 CONFLICT quando si richiede un token
        mentre uno precedente è ancora attivo lato server (il token esistente
        viene "restituito nel messaggio"): se abbiamo quel token in cache lo
        riusiamo.

        C'è però un caso in cui la cache è vuota pur essendoci un token
        ancora attivo lato server: un riavvio/reload di Home Assistant entro
        i 10 minuti di validità del token precedente. La nuova istanza di
        DuretiApiClient parte senza memoria di quel token, quindi non ha
        nulla da riusare quando arriva il 409. In questo caso specifico
        ritentiamo un numero limitato di volte con una breve attesa: spesso
        il conflitto si risolve comunque in pochi secondi/minuti, e non
        vogliamo bloccare il setup dell'integrazione abbastanza a lungo da
        far scattare il timeout di Home Assistant (che è sull'ordine dei
        minuti, non dei secondi).
        """
        now = time.monotonic()
        if not force_refresh and self._token and now < self._token_expiry:
            return self._token

        payload = {"clientId": self._client_id, "secretId": self._secret_id}

        for tentativo in range(1, RIAVVIO_RETRY_ATTEMPTS + 1):
            try:
                _log_richiesta(URL_REQUEST_TOKEN, payload, None)
                async with self._session.post(URL_REQUEST_TOKEN, json=payload) as resp:
                    data = await self._json_or_raise(resp)
                break
            except DuretiConflictError:
                if self._token:
                    _LOGGER.debug(
                        "requestToken in conflitto (409): il token precedente è ancora attivo, lo riuso"
                    )
                    # Non conosciamo la scadenza esatta residua: prudenzialmente
                    # concediamo un margine più ampio prima del prossimo tentativo.
                    self._token_expiry = now + TOKEN_SAFETY_MARGIN_SECONDS
                    return self._token

                if tentativo == RIAVVIO_RETRY_ATTEMPTS:
                    _LOGGER.warning(
                        "requestToken in conflitto (409) senza token in cache dopo %d tentativi: "
                        "probabile riavvio/reload avvenuto entro i 10 minuti di validità di un "
                        "token precedente. Non è un blocco del WAF: si risolverà da solo appena "
                        "quel token scade lato server (Home Assistant ritenterà automaticamente).",
                        RIAVVIO_RETRY_ATTEMPTS,
                    )
                    raise
                _LOGGER.debug(
                    "requestToken in conflitto (409) senza token in cache (probabile riavvio "
                    "recente), ritento tra %ds (tentativo %d/%d)",
                    RIAVVIO_RETRY_WAIT_SECONDS,
                    tentativo,
                    RIAVVIO_RETRY_ATTEMPTS,
                )
                await asyncio.sleep(RIAVVIO_RETRY_WAIT_SECONDS)

        if data.get("esito") != ESITO_OK:
            raise DuretiAuthError(data.get("message") or "Autenticazione fallita")

        token = data.get("token")
        if not token:
            raise DuretiAuthError("Token mancante nella risposta")

        self._token = token
        self._token_expiry = now + TOKEN_VALIDITY_SECONDS - TOKEN_SAFETY_MARGIN_SECONDS
        return token

    async def _json_or_raise(self, resp: aiohttp.ClientResponse) -> dict:
        """Legge il body JSON (anche in caso di errore: il manuale conferma che
        tutte le risposte, comprese quelle 4xx, hanno body JSON con esito/message)
        e solleva l'eccezione specifica in base allo status HTTP documentato.

        Legge sempre prima il testo grezzo e lo logga (a DEBUG): quando il WAF
        interviene la risposta non è JSON ma HTML, e il testo esatto con il
        Support ID è l'informazione utile per il ticket di assistenza.
        """
        import json as _json

        testo = await resp.text()
        _log_risposta(str(resp.url), resp.status, resp.headers, testo)

        try:
            data = _json.loads(testo)
        except ValueError:
            raise DuretiApiError(
                f"Risposta non JSON (HTTP {resp.status}): {testo[:300]}", http_status=resp.status
            )

        if resp.status == 200:
            return data

        message = data.get("message") or f"HTTP {resp.status}"
        if resp.status == 401:
            raise DuretiAuthError(message, http_status=401, data=data)
        if resp.status == 409:
            raise DuretiConflictError(message, http_status=409, data=data)
        if resp.status == 429:
            raise DuretiRateLimitError(message, http_status=429, data=data)
        if resp.status == 404:
            raise DuretiNotFoundError(message, http_status=404, data=data)
        if resp.status == 400:
            raise DuretiValidationError(message, http_status=400, data=data)
        raise DuretiApiError(message, http_status=resp.status, data=data)

    async def request_export(
        self,
        data_da: date,
        data_a: date,
        pods: list[dict],
        mode: str = MODE_CURVE,
    ) -> str:
        """Prenota l'estrazione e restituisce il ticket.

        pods: lista di dict {"pod": <codice POD/PDR>, "df": <dato fiscale associato>}
        """
        if len(pods) > MAX_SUPPLY_POINTS_PER_REQUEST:
            raise DuretiApiError(
                f"Troppi POD/PDR in una richiesta ({len(pods)}), "
                f"il limite dichiarato dal manuale è {MAX_SUPPLY_POINTS_PER_REQUEST}"
            )
        if (data_a.year - data_da.year) * 12 + (data_a.month - data_da.month) > MAX_DATE_RANGE_MONTHS:
            raise DuretiApiError(
                f"Range di date troppo ampio, il limite dichiarato dal manuale è "
                f"{MAX_DATE_RANGE_MONTHS} mesi"
            )

        token = await self._get_token()
        body = {
            # Confermato via test reale: il formato corretto è yyyy-mm-dd
            # (l'esempio JSON del manuale), non dd-mm-yyyy come diceva il
            # testo descrittivo.
            "dataDa": data_da.strftime("%Y-%m-%d"),
            "dataA": data_a.strftime("%Y-%m-%d"),
            "mode": mode,
            "supplyPoints": [{"supplyPoint": p["pod"], "df": p["df"]} for p in pods],
        }
        headers = {"Authorization": token}
        try:
            _log_richiesta(URL_REQUEST_EXPORT, body, headers)
            async with self._session.post(URL_REQUEST_EXPORT, json=body, headers=headers) as resp:
                data = await self._json_or_raise(resp)
        except DuretiValidationError as err:
            # Il manuale segnala un 400 specifico quando esiste già
            # un'elaborazione per gli stessi dati: il ticket collegato viene
            # comunicato nel messaggio, in un formato non specificato.
            # Proviamo a estrarlo con una regex permissiva; se non troviamo
            # nulla, rilanciamo l'errore originale.
            match = re.search(r"[A-Za-z0-9+/]{16,}", err.args[0])
            if match:
                _LOGGER.info(
                    "requestExport: elaborazione già presente, riuso il ticket dal messaggio (%s)",
                    err.args[0],
                )
                return match.group(0)
            raise

        if data.get("esito") != ESITO_OK:
            raise DuretiApiError(data.get("message") or "requestExport fallita")

        ticket = data.get("ticket")
        if not ticket:
            raise DuretiApiError("Ticket mancante nella risposta di requestExport")
        return ticket

    async def request_result(self, ticket: str) -> bytes:
        """Esegue il polling su requestResult finché il file non è pronto.

        Confermato via test reale: mentre il job è in coda la risposta è
        {"esito": 1, "message": "Il file non è ancora disponibile"} - lo
        trattiamo quindi come "non pronto ancora" e ritentiamo ogni
        RESULT_POLL_INTERVAL_SECONDS (default 10 minuti) fino a
        RESULT_POLL_MAX_ATTEMPTS (default ~6 ore totali). Il token, valido
        10 minuti, viene rinnovato automaticamente ad ogni tentativo.
        """
        for attempt in range(1, RESULT_POLL_MAX_ATTEMPTS + 1):
            # Il token dura 10 minuti quanto l'intervallo di polling: lo
            # richiediamo/rinnoviamo ad ogni tentativo (_get_token usa la
            # cache se ancora valido, quindi qui è economico).
            token = await self._get_token()
            headers = {"Authorization": token}
            body = {"ticket": ticket}

            try:
                _log_richiesta(URL_REQUEST_RESULT, body, headers)
                async with self._session.post(URL_REQUEST_RESULT, json=body, headers=headers) as resp:
                    data = await self._json_or_raise(resp)
            except DuretiAuthError:
                # Token rifiutato nonostante la cache: forziamo un rinnovo e
                # ritentiamo subito, senza consumare un intero intervallo di attesa.
                _LOGGER.debug("requestResult: token rifiutato (401), forzo il rinnovo e ritento")
                await self._get_token(force_refresh=True)
                continue
            except DuretiRateLimitError as err:
                _LOGGER.warning("requestResult: limite richieste raggiunto (429), attendo di più (%s)", err)
                await asyncio.sleep(RESULT_POLL_INTERVAL_SECONDS * 2)
                continue
            except DuretiNotFoundError as err:
                # Ticket non collegato a nessun dato: errore permanente, non ha
                # senso continuare il polling.
                # Ticket presente ma non collegato a nessun dato: errore
                # permanente, non ha senso continuare il polling. Rilanciamo il
                # tipo originale (non la classe base) così il chiamante può
                # distinguere "ticket davvero invalido" - l'unico caso in cui
                # ha senso scartarlo - da un errore transitorio.
                raise DuretiNotFoundError(
                    f"Ticket non valido o non trovato: {err}",
                    http_status=404,
                    data=getattr(err, "data", None),
                ) from err
            except DuretiApiError as err:
                if type(err) is not DuretiApiError:
                    # Sottoclasse nota non gestita sopra (es. 400 di
                    # validazione): è un errore applicativo vero, non c'entra
                    # il WAF, non ha senso continuare a ritentare.
                    raise
                # DuretiApiError "puro" (nessuna sottoclasse): tipicamente una
                # risposta non-JSON, il sintomo del blocco WAF (vedi
                # _json_or_raise). Non è un segnale sullo stato del ticket -
                # Duereti potrebbe benissimo starlo ancora processando
                # normalmente. Lo trattiamo come "non ancora pronto" e
                # continuiamo il polling con lo STESSO ticket, invece di
                # scartarlo e perdere il lavoro già fatto da Duereti.
                _LOGGER.warning(
                    "requestResult tentativo %s/%s: possibile blocco WAF (%s), continuo il "
                    "polling con lo stesso ticket invece di scartarlo",
                    attempt,
                    RESULT_POLL_MAX_ATTEMPTS,
                    err,
                )
                await asyncio.sleep(RESULT_POLL_INTERVAL_SECONDS)
                continue

            if data.get("esito") == ESITO_OK and data.get("File"):
                return self._decode_file_field(data["File"])

            _LOGGER.debug(
                "requestResult tentativo %s/%s: non ancora pronto (%s)",
                attempt,
                RESULT_POLL_MAX_ATTEMPTS,
                data.get("message"),
            )
            await asyncio.sleep(RESULT_POLL_INTERVAL_SECONDS)

        raise DuretiApiError(
            f"File non disponibile dopo {RESULT_POLL_MAX_ATTEMPTS} tentativi (ticket={ticket})"
        )

    @staticmethod
    def _decode_file_field(file_field: str) -> bytes:
        """Decodifica il campo 'File' della risposta.

        Confermato dal manuale: 'File' è sempre il Base64 dello zip
        (contenente .csv per CURVE o .xlsx per LETTURE).
        """
        try:
            return base64.b64decode(file_field)
        except Exception as err:  # noqa: BLE001
            raise DuretiApiError("Impossibile decodificare il campo File come base64") from err


@dataclass
class LetturaRiga:
    """Una riga di lettura periodica (mode=LETTURE), formato confermato via test reale."""

    data_lettura: "datetime"  # noqa: F821
    tipo_lettura: str  # es. "SALDO"
    valore: float
    tipologia_misura: str  # es. "Energia attiva F3"
    matricola_contatore: str
    tipo_misuratore: str
    energia: str  # es. "Energia attiva", "Energia reattiva", "Potenza massima"
    fascia: str  # es. "F1", "F2", "F3"


def parse_letture_zip(zip_bytes: bytes) -> dict[str, list[LetturaRiga]]:
    """Estrae e parsa i file .xlsx dentro lo zip restituito da Duereti per mode=LETTURE.

    Formato CONFERMATO via test reale (non più ipotesi):
    colonne "Data lettura", "Tipo lettura", "Lettura", "Tipologia Misura",
    "Matricola Contatore", "Tipo Misuratore", "Energia", "Fascia". Il foglio
    Excel è nominato come il POD/PDR richiesto, "Lettura" è il valore
    cumulativo (SALDO) del contatore a quella data, non un delta.

    NOTA: questa funzione non è collegata al coordinator dell'integrazione,
    che per scelta usa solo mode=CURVE (dati a intervalli, più adatti alla
    Energy Dashboard). È qui come riferimento/utility nel caso in futuro
    serva anche l'import delle letture periodiche. Richiede 'openpyxl'
    (import lazy per non renderlo una dipendenza hard dell'integrazione).
    """
    import openpyxl
    from datetime import datetime

    risultati: dict[str, list[LetturaRiga]] = {}

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            if not name.lower().endswith(".xlsx"):
                continue
            with zf.open(name) as f:
                wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)

            for sheet_name in wb.sheetnames:
                # Il foglio è nominato come il POD/PDR
                pod = sheet_name
                ws = wb[sheet_name]
                rows = ws.iter_rows(values_only=True)
                header = next(rows, None)
                if not header:
                    continue
                idx = {col: i for i, col in enumerate(header)}

                for row in rows:
                    try:
                        data_lettura = datetime.strptime(row[idx["Data lettura"]], "%d/%m/%Y")
                        valore = float(str(row[idx["Lettura"]]).replace(",", "."))
                    except (KeyError, ValueError, TypeError, IndexError):
                        _LOGGER.warning("Riga non riconosciuta nel foglio %s: %s", sheet_name, row)
                        continue

                    risultati.setdefault(pod, []).append(
                        LetturaRiga(
                            data_lettura=data_lettura,
                            tipo_lettura=row[idx["Tipo lettura"]],
                            valore=valore,
                            tipologia_misura=row[idx["Tipologia Misura"]],
                            matricola_contatore=row[idx["Matricola Contatore"]],
                            tipo_misuratore=row[idx["Tipo Misuratore"]],
                            energia=row[idx["Energia"]],
                            fascia=row[idx["Fascia"]],
                        )
                    )

    return risultati


def parse_curve_zip(zip_bytes: bytes) -> dict[str, RisultatoLetture]:
    """Estrae e parsa i file .csv dentro lo zip restituito da Duereti per mode=CURVE.

    Formato CONFERMATO via test reale (non più ipotesi), CSV separato da ';':
        POD;DATA;ORA;FL_ORA_LEGALE;ATTIVA_PRELEVATA;ATTIVA_IMMESSA;
        REATTIVA_CAPACITIVA_IMMESSA;REATTIVA_CAPACITIVA_PRELEVATA;
        REATTIVA_INDUTTIVA_IMMESSA;REATTIVA_INDUTTIVA_PRELEVATA;
        PICCO_PRELEVATA;CONSUMO_PICCO_IMMESSA;TIPO_DATO;

    Esempio riga:
        IT001E00000001;20260501;000000;2;0,025;;;;;0,001;5,004;;E;

    Note sul formato:
    - DATA (yyyymmdd) + ORA (hhmmss) = timestamp, a intervalli di 15 minuti
    - i decimali usano la virgola, non il punto
    - ATTIVA_PRELEVATA è energia CONSUMATA nell'intervallo (già un delta per
      i 15 minuti, non un valore cumulativo) - è il campo che importiamo
    - ATTIVA_IMMESSA è energia immessa in rete (produzione, es. fotovoltaico):
      non ancora importata da questa funzione, TODO se in futuro serve
      gestire anche la produzione
    - i campi reattivi/picco non vengono importati
    """
    from datetime import datetime

    risultati: dict[str, RisultatoLetture] = {}

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for name in zf.namelist():
            if not name.lower().endswith((".csv", ".txt")):
                continue
            with zf.open(name) as f:
                raw = f.read().decode("utf-8-sig", errors="replace")

            reader = csv.DictReader(io.StringIO(raw), delimiter=";")

            for row in reader:
                pod = row.get("POD")
                data_raw = row.get("DATA")
                ora_raw = row.get("ORA")
                valore_raw = row.get("ATTIVA_PRELEVATA")

                if not (pod and data_raw and ora_raw):
                    continue

                try:
                    ts = datetime.strptime(f"{data_raw}{ora_raw}", "%Y%m%d%H%M%S")
                except ValueError:
                    _LOGGER.warning("Timestamp non parsabile: DATA=%s ORA=%s", data_raw, ora_raw)
                    continue

                if not valore_raw:
                    # Intervallo senza consumo registrato (es. solo produzione,
                    # o dato mancante): saltiamo, non è un errore.
                    continue

                try:
                    valore = float(valore_raw.replace(",", "."))
                except ValueError:
                    _LOGGER.warning("Valore ATTIVA_PRELEVATA non parsabile: %s", valore_raw)
                    continue

                risultati.setdefault(pod, RisultatoLetture(pod=pod)).punti.append(
                    CurvaPunto(
                        timestamp=ts,
                        valore_kwh=valore,
                        ora_legale=(row.get("FL_ORA_LEGALE") or "").strip() or None,
                    )
                )

    return risultati
